"""
工具函数模块
包含文件处理、图片处理、Excel导出等功能
"""

import os
from pathlib import Path
from datetime import datetime
# 延迟导入重型库，减小打包体积
# import fitz  # PyMuPDF - 改为按需导入
from PIL import Image
# import openpyxl - 改为按需导入
# from openpyxl.styles import Font, Alignment, Border, Side - 改为按需导入
from config import Config


class FileUtils:
    """文件处理工具类"""
    
    @staticmethod
    def is_supported_file(file_path):
        """检查文件是否为支持的格式"""
        ext = Path(file_path).suffix.lower()
        return ext in Config.get_all_supported_formats()
    
    @staticmethod
    def is_image_file(file_path):
        """检查是否为图片文件"""
        ext = Path(file_path).suffix.lower()
        return ext in Config.SUPPORTED_IMAGE_FORMATS
    
    @staticmethod
    def is_pdf_file(file_path):
        """检查是否为PDF文件"""
        ext = Path(file_path).suffix.lower()
        return ext in Config.SUPPORTED_PDF_FORMAT
    
    @staticmethod
    def get_files_from_folder(folder_path, recursive=True):
        """
        从文件夹获取所有支持的文件
        :param folder_path: 文件夹路径
        :param recursive: 是否递归搜索子文件夹
        :return: 文件路径列表
        """
        files = []
        
        if recursive:
            for root, dirs, filenames in os.walk(folder_path):
                for filename in filenames:
                    full_path = os.path.join(root, filename)
                    if FileUtils.is_supported_file(full_path):
                        files.append(full_path)
        else:
            for filename in os.listdir(folder_path):
                full_path = os.path.join(folder_path, filename)
                if os.path.isfile(full_path) and FileUtils.is_supported_file(full_path):
                    files.append(full_path)
        
        return sorted(files)
    
    @staticmethod
    def clean_filename(filename):
        """
        清理文件名中的非法字符
        :param filename: 原始文件名
        :return: 清理后的文件名
        """
        # 移除非法字符
        cleaned = "".join(c for c in filename if c not in Config.RENAME_INVALID_CHARS)
        
        # 移除首尾空格
        cleaned = cleaned.strip()
        
        # 如果文件名为空，使用默认名称
        if not cleaned:
            cleaned = "unnamed"
        
        return cleaned
    
    @staticmethod
    def get_unique_filename(directory, base_name, extension):
        """
        生成唯一的文件名（避免重名）- Windows风格括号序号
        :param directory: 目录路径
        :param base_name: 基础文件名
        :param extension: 文件扩展名
        :return: 唯一的完整文件路径
        """
        full_path = os.path.join(directory, base_name + extension)
        
        if not os.path.exists(full_path):
            return full_path
        
        # 如果文件已存在，添加括号序号后缀 (Windows风格)
        counter = 1
        while True:
            new_name = f"{base_name}({counter}){extension}"
            full_path = os.path.join(directory, new_name)
            if not os.path.exists(full_path):
                return full_path
            counter += 1


class ImageUtils:
    """图片处理工具类"""
    
    @staticmethod
    def load_image(file_path):
        """
        加载图片文件
        :param file_path: 文件路径
        :return: PIL Image对象
        """
        if FileUtils.is_pdf_file(file_path):
            return ImageUtils.pdf_to_image(file_path)
        else:
            img = Image.open(file_path)
            # 转换为RGB模式
            if img.mode != 'RGB':
                img = img.convert('RGB')
            return img
    
    @staticmethod
    def pdf_to_image(pdf_path, page_num=0, zoom=None):
        """
        将PDF转换为图片
        :param pdf_path: PDF文件路径
        :param page_num: 页码（从0开始）
        :param zoom: 缩放因子
        :return: PIL Image对象
        """
        # 按需导入PyMuPDF，只有处理PDF时才加载
        import fitz
        
        if zoom is None:
            zoom = Config.PDF_ZOOM_FACTOR
        
        doc = fitz.open(pdf_path)
        
        # 检查页码是否有效
        if page_num >= len(doc):
            page_num = 0
        
        page = doc[page_num]
        
        # 设置缩放比例
        mat = fitz.Matrix(zoom, zoom)
        pix = page.get_pixmap(matrix=mat)
        
        # 转换为PIL Image
        img = Image.frombytes("RGB", [pix.width, pix.height], pix.samples)
        doc.close()
        
        return img
    
    @staticmethod
    def calculate_display_size(img_width, img_height, canvas_width, canvas_height, max_scale=None):
        """
        计算图片在画布上的显示尺寸
        :param img_width: 图片宽度
        :param img_height: 图片高度
        :param canvas_width: 画布宽度
        :param canvas_height: 画布高度
        :param max_scale: 最大缩放比例
        :return: (新宽度, 新高度, 缩放因子)
        """
        if max_scale is None:
            max_scale = Config.MAX_DISPLAY_SCALE
        
        # 计算缩放比例
        scale_w = canvas_width / img_width
        scale_h = canvas_height / img_height
        scale_factor = min(scale_w, scale_h, max_scale)
        
        # 计算新尺寸
        new_width = int(img_width * scale_factor)
        new_height = int(img_height * scale_factor)
        
        return new_width, new_height, scale_factor
    
    @staticmethod
    def resize_image(image, new_width, new_height):
        """
        调整图片尺寸
        :param image: PIL Image对象
        :param new_width: 新宽度
        :param new_height: 新高度
        :return: 调整后的PIL Image对象
        """
        return image.resize((new_width, new_height), Image.Resampling.LANCZOS)


class ExcelExporter:
    """Excel导出工具类"""
    
    @staticmethod
    def load_existing_data(file_path):
        """
        读取现有Excel文件的数据
        :param file_path: Excel文件路径
        :return: (existing_data_rows, max_existing_rects) 或 (None, 0)
        """
        if not os.path.exists(file_path):
            return None, 0
        
        try:
            # 按需导入openpyxl
            import openpyxl
            
            # 加载现有工作簿
            wb = openpyxl.load_workbook(file_path)
            ws = wb.active
            
            # 读取表头，解析区域列数量
            headers = []
            for cell in ws[1]:
                if cell.value:
                    headers.append(cell.value)
            
            # 计算现有的区域列数量
            # 表头格式: ["序号", "文件名", "文件路径", "识别时间", "状态", "区域1", "区域2", ...]
            fixed_columns = 5  # 前5列是固定的
            max_existing_rects = max(0, len(headers) - fixed_columns)
            
            # 读取所有数据行（跳过表头）
            existing_data_rows = []
            for row in ws.iter_rows(min_row=2, values_only=True):
                # 过滤掉空行
                if row and any(cell is not None for cell in row):
                    existing_data_rows.append(list(row))
            
            wb.close()
            
            return existing_data_rows, max_existing_rects
            
        except Exception as e:
            print(f"读取现有Excel文件失败: {e}")
            return None, 0
    
    @staticmethod
    def export_results(ocr_results, save_path, append_mode=False):
        """
        导出OCR识别结果到Excel
        :param ocr_results: 识别结果字典 {file_path: {rects: [], status: ""}}
        :param save_path: 保存路径
        :param append_mode: 是否追加模式（True=追加到现有文件，False=新建文件）
        :return: 是否成功
        """
        try:
            # 按需导入openpyxl，只有导出Excel时才加载
            import openpyxl
            from openpyxl.styles import Font, Alignment, Border, Side
            
            # 处理追加模式和新建模式
            existing_data_rows = []
            max_existing_rects = 0
            start_index = 1  # 新数据的起始序号
            
            if append_mode and os.path.exists(save_path):
                # 追加模式：读取现有数据
                existing_data_rows, max_existing_rects = ExcelExporter.load_existing_data(save_path)
                if existing_data_rows is None:
                    # 读取失败，提示错误
                    print(f"追加模式下读取现有文件失败，将创建新文件")
                    existing_data_rows = []
                    max_existing_rects = 0
                else:
                    # 计算新数据的起始序号（从现有数据的最大序号+1开始）
                    if existing_data_rows:
                        # 第一列是序号
                        max_existing_index = max(row[0] for row in existing_data_rows if row and row[0] is not None)
                        start_index = max_existing_index + 1
            elif not append_mode and os.path.exists(save_path):
                # 新建模式：文件已存在，生成唯一文件名
                import os.path as path
                directory = path.dirname(save_path)
                filename = path.basename(save_path)
                base_name, extension = path.splitext(filename)
                save_path = FileUtils.get_unique_filename(directory, base_name, extension)
            
            # 获取新数据的最大区域数量
            max_new_rects = 0
            for result in ocr_results.values():
                rects = result.get("rects", [])
                if len(rects) > max_new_rects:
                    max_new_rects = len(rects)
            
            # 计算全局最大区域数量
            max_rects = max(max_existing_rects, max_new_rects)
            
            # 调试信息
            if append_mode:
                print(f"[Excel追加] 旧数据区域数: {max_existing_rects}, 新数据区域数: {max_new_rects}, 最终区域数: {max_rects}")
            
            # 创建工作簿
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "OCR识别结果"
            
            # 设置表头
            headers = ["序号", "文件名", "文件路径", "识别时间", "状态"]
            for i in range(max_rects):
                headers.append(f"区域{i+1}")
            
            ws.append(headers)
            
            # 设置表头样式
            header_font = Font(bold=True, size=Config.EXCEL_HEADER_FONT_SIZE)
            header_alignment = Alignment(horizontal='center', vertical='center')
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            for cell in ws[1]:
                cell.font = header_font
                cell.alignment = header_alignment
                cell.border = border
            
            # 先写入现有数据（如果是追加模式）
            if existing_data_rows:
                print(f"[Excel追加] 开始写入 {len(existing_data_rows)} 行旧数据")
                for idx, row_data in enumerate(existing_data_rows, 1):
                    original_len = len(row_data)
                    
                    # 补齐区域列（如果新数据有更多区域）
                    while len(row_data) < len(headers):
                        row_data.append("")
                    
                    # 截断多余的列（如果旧数据列数超过新表头）
                    if len(row_data) > len(headers):
                        row_data = row_data[:len(headers)]
                    
                    ws.append(row_data)
                    
                    # 调试：显示第一行的处理情况
                    if idx == 1:
                        print(f"[Excel追加] 第1行旧数据: 原长度={original_len}, 补齐后={len(row_data)}, 表头列数={len(headers)}")
                    
                    # 设置边框
                    for cell in ws[ws.max_row]:
                        cell.border = border
                
                print(f"[Excel追加] 旧数据写入完成")
            
            # 填充新数据
            print(f"[Excel追加] 开始写入 {len(ocr_results)} 行新数据")
            first_new_row = True
            for idx, (file_path, result) in enumerate(ocr_results.items(), start_index):
                row_data = [
                    idx,
                    os.path.basename(file_path),
                    file_path,
                    datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                    result.get("status", "未知")
                ]
                
                # 添加各区域识别文本
                rects = result.get("rects", [])
                for rect in rects:
                    row_data.append(rect.text if hasattr(rect, 'text') else "")
                
                # 调试：显示第一行新数据的区域数量
                if first_new_row:
                    print(f"[Excel追加] 第1行新数据: 区域数={len(rects)}, 数据列数={len(row_data)-5}")
                    first_new_row = False
                
                # 补齐区域列
                while len(row_data) < len(headers):
                    row_data.append("")
                
                ws.append(row_data)
                
                # 设置边框
                for cell in ws[ws.max_row]:
                    cell.border = border
            
            print(f"[Excel追加] 新数据写入完成")
            
            # 自动调整列宽
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                
                for cell in column:
                    try:
                        if cell.value:
                            cell_length = len(str(cell.value))
                            if cell_length > max_length:
                                max_length = cell_length
                    except:
                        pass
                
                adjusted_width = min(max_length + 2, Config.EXCEL_MAX_COLUMN_WIDTH)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # 保存
            wb.save(save_path)
            return True
            
        except Exception as e:
            print(f"导出Excel失败: {e}")
            return False
    
    @staticmethod
    def export_batch_results(file_results, save_path):
        """
        批量导出识别结果
        :param file_results: 文件结果列表
        :param save_path: 保存路径
        :return: 是否成功
        """
        # 转换为标准格式
        ocr_results = {}
        for file_path, rects, status in file_results:
            ocr_results[file_path] = {
                "rects": rects,
                "status": status
            }
        
        return ExcelExporter.export_results(ocr_results, save_path)
